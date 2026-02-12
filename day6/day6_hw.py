import os
import re
import uuid
import torch
import requests
import gc
from qdrant_client import QdrantClient, models
from transformers import AutoTokenizer, AutoModelForCausalLM
from openai import OpenAI
from deepeval.models import DeepEvalBaseLLM
from deepeval.metrics import (
    FaithfulnessMetric,
    AnswerRelevancyMetric,
    ContextualRecallMetric,
    ContextualPrecisionMetric,
    ContextualRelevancyMetric,
)
from deepeval.test_case import LLMTestCase
from deepeval import evaluate
from deepeval.evaluate.configs import DisplayConfig, AsyncConfig

# ==========================================
# 【系統安全設定】
# ==========================================
# 即使有 GPU，CPU 仍負責資料處理。限制使用 4 核心，確保電腦不卡頓。
torch.set_num_threads(4)
os.environ["OMP_NUM_THREADS"] = "4"

# 顯示目前的 PyTorch 狀態
print(f"PyTorch Version: {torch.__version__}")
print(f"CUDA Available: {torch.cuda.is_available()}")
if torch.cuda.is_available():
    print(f"GPU Device: {torch.cuda.get_device_name(0)}")
    # 預先清理 GPU 記憶體
    torch.cuda.empty_cache()
    gc.collect()

# ==========================================

class LlamaCppModel(DeepEvalBaseLLM):
    def __init__(
        self,
        base_url="https://ws-02.wade0426.me/v1",
        model_name="gpt-4o"
    ):
        self.base_url = base_url
        self.model_name = model_name
        
    def load_model(self):
        return OpenAI(
            api_key="NoNeed",
            base_url=self.base_url,
            timeout=300.0
        )
    
    def generate(self, prompt: str) -> str:
        client = self.load_model()
        response = client.chat.completions.create(
            model=self.model_name,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.7,
        )
        return response.choices[0].message.content
    
    async def a_generate(self, prompt: str) -> str:
        return self.generate(prompt)
    
    def get_model_name(self):
        return f"Llama.cpp ({self.model_name})"


# ========== 設定 ==========
QDRANT_URL = "http://localhost:6333"
COLLECTION_NAME = "day6_hw"
MODEL_PATH = r"c:\Users\berto\Desktop\lab\homework\test\day6\~\Models\Qwen3-Reranker-0.6B"
EMBEDDING_API_URL = "https://ws-04.wade0426.me/embed"
DATA_FILE = os.path.join(os.path.dirname(__file__), "hw_data", "qa_data.txt")

# ========== 1. 初始化 Qdrant ==========
client = QdrantClient(url=QDRANT_URL)

# ========== 2. 智慧載入 Reranker 模型 ==========
print(f"\n正在載入 Reranker 模型...")

# 預設參數
reranker_device = "cpu"
use_fp16 = False

# 嘗試使用 GPU
if torch.cuda.is_available():
    try:
        print("嘗試將模型載入 GPU (MX350) 並啟用 FP16 優化...")
        
        # 這裡不直接載入 model，先載入 tokenizer
        reranker_tokenizer = AutoTokenizer.from_pretrained(
            MODEL_PATH,
            local_files_only=True,
            trust_remote_code=True,
            padding_side='left'
        )

        # 嘗試載入模型到 GPU，強制使用 FP16 (省一半記憶體)
        reranker_model = AutoModelForCausalLM.from_pretrained(
            MODEL_PATH,
            local_files_only=True,
            trust_remote_code=True,
            device_map="cuda",          # 指定 GPU
            torch_dtype=torch.float16,  # 【關鍵】使用 FP16 半精度
            low_cpu_mem_usage=True
        ).eval()
        
        reranker_device = "cuda"
        use_fp16 = True
        print("✅ 成功！模型已在 GPU 上運行 (FP16 模式)。")
        print("注意：MX350 只有 2GB VRAM，請勿開啟大量背景程式。")

    except Exception as e:
        print(f"❌ GPU 載入失敗 (可能是記憶體不足): {e}")
        print("🔄 正在切換回 CPU 模式 (不用擔心，這很正常)...")
        
        # 清理失敗的殘留記憶體
        if 'reranker_model' in locals(): del reranker_model
        torch.cuda.empty_cache()
        gc.collect()
        
        # CPU Fallback
        reranker_model = AutoModelForCausalLM.from_pretrained(
            MODEL_PATH,
            local_files_only=True,
            trust_remote_code=True,
            device_map="cpu",
            torch_dtype=torch.float32
        ).eval()
        reranker_device = "cpu"
else:
    print("未偵測到 GPU，直接使用 CPU 模式。")
    # 如果沒 GPU，需要補載入 tokenizer
    if 'reranker_tokenizer' not in locals():
        reranker_tokenizer = AutoTokenizer.from_pretrained(
            MODEL_PATH,
            local_files_only=True,
            trust_remote_code=True,
            padding_side='left'
        )
    reranker_model = AutoModelForCausalLM.from_pretrained(
            MODEL_PATH,
            local_files_only=True,
            trust_remote_code=True,
            device_map="cpu",
            torch_dtype=torch.float32
        ).eval()


# 獲取 token IDs
token_false_id = reranker_tokenizer.convert_tokens_to_ids("no")
token_true_id = reranker_tokenizer.convert_tokens_to_ids("yes")

# 【重要】設定最大長度
# MX350 記憶體小，設太大必當機。1024 對大多數 QA 足夠。
# 如果還是爆顯存，請將此數字改為 512
max_reranker_length = 1024 
print(f"設定 Reranker 最大長度限制: {max_reranker_length} tokens")

prefix = "<|im_start|>system\nJudge whether the Document meets the requirements based on the Query and the Instruct provided. Note that the answer can only be \"yes\" or \"no\".<|im_end|>\n<|im_start|>user\n"
suffix = "<|im_end|>\n<|im_start|>assistant\n<think>\n\n</think>\n\n"
prefix_tokens = reranker_tokenizer.encode(prefix, add_special_tokens=False)
suffix_tokens = reranker_tokenizer.encode(suffix, add_special_tokens=False)


# ========== 輔助函數 ==========

def get_embeddings(texts: list) -> list:
    data = {"texts": texts, "normalize": True, "batch_size": 32}
    try:
        response = requests.post(EMBEDDING_API_URL, json=data)
        if response.status_code == 200:
            return response.json()['embeddings']
        return []
    except Exception as e:
        print(f"Embedding Error: {e}")
        return []

def format_instruction(instruction, query, doc):
    if instruction is None: instruction = '根據查詢檢索相關文件'
    return f"<Instruct>: {instruction}\n<Query>: {query}\n<Document>: {doc}"

def process_inputs(pairs):
    inputs = reranker_tokenizer(
        pairs,
        padding=False,
        truncation='longest_first',
        return_attention_mask=False,
        max_length=max_reranker_length - len(prefix_tokens) - len(suffix_tokens)
    )
    for i, ele in enumerate(inputs['input_ids']):
        inputs['input_ids'][i] = prefix_tokens + ele + suffix_tokens
    
    inputs = reranker_tokenizer.pad(
        inputs,
        padding=True,
        return_tensors="pt",
        max_length=max_reranker_length
    )
    
    # 將輸入移動到與模型相同的設備 (GPU or CPU)
    for key in inputs:
        inputs[key] = inputs[key].to(reranker_model.device)
    
    return inputs

@torch.no_grad()
def compute_logits(inputs):
    # 計算 logits
    batch_scores = reranker_model(**inputs).logits[:, -1, :]
    true_vector = batch_scores[:, token_true_id]
    false_vector = batch_scores[:, token_false_id]
    batch_scores = torch.stack([false_vector, true_vector], dim=1)
    batch_scores = torch.nn.functional.log_softmax(batch_scores, dim=1)
    return batch_scores[:, 1].exp().tolist()

def rerank_documents(query, documents, task_instruction=None):
    if task_instruction is None: task_instruction = '根據查詢檢索相關文件'
    texts = [doc[0] for doc in documents]
    pairs = [format_instruction(task_instruction, query, text) for text in texts]

    try:
        # MX350 批次處理能力極弱，我們強制將 batch_size 設為 1
        # 這會比一次處理慢一點，但能保證不爆顯存
        scores = []
        batch_size = 1 
        
        for i in range(0, len(pairs), batch_size):
            batch_pairs = pairs[i : i + batch_size]
            inputs = process_inputs(batch_pairs)
            batch_scores = compute_logits(inputs)
            scores.extend(batch_scores)
            
            # 如果在 GPU 上，跑完一筆就清一下垃圾
            if reranker_device == "cuda":
                del inputs
                torch.cuda.empty_cache()

    except RuntimeError as e:
        if "out of memory" in str(e):
            print("⚠️ 顯卡記憶體不足 (OOM)，無法執行重排。跳過重排步驟。")
            torch.cuda.empty_cache()
            return [(doc, 0.0) for doc in documents]
        else:
            print(f"Rerank Error: {e}")
            return [(doc, 0.0) for doc in documents]

    doc_scores = list(zip(documents, scores))
    doc_scores.sort(key=lambda x: x[1], reverse=True)
    return doc_scores

# ========== 其他功能函數 (切分、混合搜索、LLM) ==========

def split_text_qa_aware(text):
    # (保持原樣)
    chunks = []
    lines = text.split('\n')
    date_indices = [i for i, line in enumerate(lines) if '**發布日期**' in line]
    if not date_indices: return [text]

    for idx, date_line_idx in enumerate(date_indices):
        question = lines[date_line_idx - 1].strip() if date_line_idx > 0 else ""
        date_match = re.search(r'(\d{4}/\d{2}/\d{2})', lines[date_line_idx])
        date_str = date_match.group(1) if date_match else ""
        end_boundary = date_indices[idx + 1] - 1 if idx + 1 < len(date_indices) else len(lines)
        
        content_lines = []
        source = ""
        for j in range(date_line_idx + 1, end_boundary):
            line = lines[j].strip()
            if not line: 
                content_lines.append("")
                continue
            if line.startswith('來源：'):
                source = line
                break
            content_lines.append(lines[j])
        
        while content_lines and not content_lines[-1].strip(): content_lines.pop()
        content = '\n'.join(content_lines).strip()
        
        if source: qa_unit = f"{question}\n**發布日期**: {date_str}\n{content}\n{source}"
        else: qa_unit = f"{question}\n**發布日期**: {date_str}\n{content}"
        chunks.append(qa_unit)
    return chunks

def hybrid_search_with_rerank(query: str, initial_limit: int = 10, final_limit: int = 3):
    embeddings = get_embeddings([query])
    if not embeddings: return []
    query_embedding = embeddings[0]

    try:
        response = client.query_points(
            collection_name=COLLECTION_NAME,
            prefetch=[
                models.Prefetch(query=models.Document(text=query, model="Qdrant/bm25"), using="sparse", limit=initial_limit),
                models.Prefetch(query=query_embedding, using="dense", limit=initial_limit),
            ],
            query=models.FusionQuery(fusion=models.Fusion.RRF),
            limit=initial_limit,
        )
    except Exception as e:
        print(f"Qdrant Search Error: {e}")
        return []

    candidate_docs = [(point.payload.get("text", ""), point.payload.get("source_file", "Unknown")) for point in response.points]
    if not candidate_docs: return []

    print(f"正在重排 {len(candidate_docs)} 個文件 (使用 {reranker_device.upper()})...")
    reranked_results = rerank_documents(query, candidate_docs)
    
    top_results = reranked_results[:final_limit]
    print(f"Top {final_limit} 分數: {[round(s, 3) for _, s in top_results]}")
    return top_results

from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage

llm = ChatOpenAI(
    model="/models/Qwen3-30B-A3B-Instruct-2507-FP8",
    temperature=0.2,
    base_url="https://ws-03.wade0426.me/v1",
    api_key="EMPTY",
)

def generate_answer_from_llm(query, context_parts):
    context_str = "\n\n".join(context_parts)
    prompt = f"""你是一位專業的知識庫助手。請根據以下【參考資訊】回答【用戶問題】。
### 參考資訊：
{context_str}
### 用戶問題：
{query}
請輸出回答，並在最後一行標註來源 (SOURCE: 檔名)："""
    try:
        messages = [HumanMessage(content=prompt)]
        response = llm.invoke(messages)
        return response.content.strip()
    except Exception as e:
        print(f"LLM Error: {e}")
        return "Error"

# ========== 評估流程 ==========
def evaluate_qa_with_deepeval():
    import openpyxl
    import traceback
    
    questions_file = "hw_data/day6_HW_questions.csv.xlsx"
    answers_file = "hw_data/questions_answer.csv.xlsx"
    
    print(f"\n========== DeepEval Evaluation ==========")
    
    # 設定 DeepEval 超時時間 (預設 180s 可能不夠)
    os.environ["DEEPEVAL_PER_ATTEMPT_TIMEOUT_SECONDS_OVERRIDE"] = "1200"

    try:
        wb_q = openpyxl.load_workbook(questions_file)
        ws_q = wb_q.active
        wb_a = openpyxl.load_workbook(answers_file)
        ws_a = wb_a.active
        
        headers_q = [c.value for c in ws_q[1]]
        q_col_q = headers_q.index('questions') + 1
        headers_a = [c.value for c in ws_a[1]]
        a_col_a = headers_a.index('answer') + 1 if 'answer' in headers_a else 3

        if 'answer' in headers_q: ans_col_q = headers_q.index('answer') + 1
        else: ans_col_q = len(headers_q) + 1; ws_q.cell(1, ans_col_q, 'answer')

        test_cases = []
        row_indices = []
        generated_answers = []
        
        # 定義需要評估的指標與對應欄位
        metric_cols = {
            "Faithfulness": "Faithfulness",
            "Answer Relevancy": "Answer_Relevancy",
            "Contextual Recall": "Contextual_Recall",
            "Contextual Precision": "Contextual_Precision",
            "Contextual Relevancy": "Contextual_Relevancy"
        }
        
        # 確保 Excel 中有這些欄位 (如果沒有則新增)
        col_mapping = {}
        for m_name, col_name in metric_cols.items():
            if col_name in headers_q:
                col_mapping[m_name] = headers_q.index(col_name) + 1
            else:
                new_col = len(headers_q) + 1
                ws_q.cell(row=1, column=new_col, value=col_name)
                headers_q.append(col_name)
                col_mapping[m_name] = new_col

        # 處理第 3-5 筆資料 (Excel Row 4, 5, 6)
        target_rows = range(4, 7)
        print(f"將處理 Row {min(target_rows)} 到 {max(target_rows)} 的資料...")

        for row_idx in target_rows: 
            question = ws_q.cell(row=row_idx, column=q_col_q).value
            if not question: continue
            
            ground_truth = ws_a.cell(row=row_idx, column=a_col_a).value or "無標準答案"
            print(f"\n[Row {row_idx}] Processing: {question}")
            
            results = hybrid_search_with_rerank(str(question).strip(), initial_limit=10, final_limit=4)
            retrieval_context = [txt for (txt, _), _ in results]
            context_parts = [f"[來源: {src}]\n{txt}" for (txt, src), _ in results]
            
            raw_answer = generate_answer_from_llm(question, context_parts)
            actual_output = raw_answer.split("SOURCE:")[0].strip() if "SOURCE:" in raw_answer else raw_answer
            print(f"  > Generated: {actual_output[:20]}...")
            
            # 先填入生成的回答 (覆寫或新增)
            ws_q.cell(row=row_idx, column=ans_col_q, value=actual_output)

            test_cases.append(LLMTestCase(
                input=str(question),
                actual_output=actual_output,
                expected_output=str(ground_truth),
                retrieval_context=retrieval_context
            ))
            row_indices.append(row_idx)
            generated_answers.append(actual_output)

        if not test_cases: return

        print("初始化評估模型...")
        # 評估模型還是會比較吃資源，DeepEval 內部無法控制這麼細
        # 但因為我們前面已經把 heavy loading 的 Reranker 處理完了，這裡應該還好
        judge_llm = LlamaCppModel()
        
        metrics = [
            FaithfulnessMetric(threshold=0.7, model=judge_llm, include_reason=False),
            AnswerRelevancyMetric(threshold=0.7, model=judge_llm, include_reason=False),
            ContextualRecallMetric(threshold=0.7, model=judge_llm, include_reason=False),
            ContextualPrecisionMetric(threshold=0.7, model=judge_llm, include_reason=False),
            ContextualRelevancyMetric(threshold=0.7, model=judge_llm, include_reason=False),
        ]
        
        print(f"開始評估 {len(test_cases)} 個案例...")
        test_results = evaluate(
            test_cases, 
            metrics,
            display_config=DisplayConfig(print_results=False, show_indicator=True),
            async_config=AsyncConfig(run_async=False)
        )
        print("評估完成，正在寫入結果...")

        # 將評估結果寫回 Excel
        try:
            for i, test_case_result in enumerate(test_results.test_results):
                row_idx = row_indices[i]
                for metric_data in test_case_result.metrics_data:
                    metric_name = metric_data.name
                    if metric_name in col_mapping:
                        col_idx = col_mapping[metric_name]
                        score = metric_data.score
                        ws_q.cell(row=row_idx, column=col_idx, value=score)
                        print(f"  Row {row_idx} - {metric_name}: {score:.4f}")
        except AttributeError as e:
            print(f"寫入評分時發生錯誤: {e}")
            print("test_results 結構:", dir(test_results))

        # 儲存結果
        output_file = "hw_data/day6_HW_questions_result.xlsx"
        wb_q.save(output_file)
        print(f"結果已儲存至: {output_file}")

    except Exception as e:
        print(f"Error: {e}")
        traceback.print_exc()

if __name__ == "__main__":
    evaluate_qa_with_deepeval()